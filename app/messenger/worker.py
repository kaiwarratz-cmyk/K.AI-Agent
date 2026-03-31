from __future__ import annotations

import asyncio
import contextlib
import json
import queue
import re
import threading
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import httpx

from app.audio import AudioService
from app.trace_utils import start_thread_with_trace


ReplyFn = Callable[[str, Dict[str, Any]], str]
LogFn = Callable[[str], None]


def _ensure_reply_text(text: str) -> str:
    # Versuche Tags wie <think> oder <details> zu filtern falls sie im Reintext stoeren
    import re
    res = str(text or "")
    res = re.sub(r'(?si)<think>.*?</think>', '', res)
    res = re.sub(r'(?si)<details>.*?</details>', '', res)
    return res.strip()


def _run_reply_with_timeout(
    reply_fn: ReplyFn,
    text: str,
    meta: Dict[str, Any],
    timeout_sec: int,
    log_fn: Optional[LogFn] = None,
    platform: str = "messenger",
) -> str:
    timeout = max(10, min(3600, int(timeout_sec or 600)))
    q: "queue.Queue[tuple[str, Any]]" = queue.Queue(maxsize=1)

    def _runner() -> None:
        try:
            out = reply_fn(text, meta)
            q.put(("ok", out), timeout=0.1)
        except Exception as exc:
            with contextlib.suppress(Exception):
                q.put(("err", exc), timeout=0.1)

    try:
        start_thread_with_trace(_runner, name=f"K.AI-{platform}-reply")
    except Exception:
        t = threading.Thread(target=_runner, daemon=True, name=f"K.AI-{platform}-reply")
        t.start()
    try:
        status, payload = q.get(timeout=timeout)
    except queue.Empty:
        if callable(log_fn):
            with contextlib.suppress(Exception):
                log_fn(f"{platform}: reply timeout nach {timeout}s")
        return (
            f"⚠️ Die Verarbeitung dauerte laenger als {timeout}s und wurde fuer den Messenger abgebrochen. "
            "Bitte sende die Anfrage erneut oder konkretisiere sie."
        )
    if status == "err":
        return f"Fehler bei Verarbeitung: {payload}"
    return str(payload or "")


def _workspace_root(raw_workspace: str) -> Path:
    p = Path(str(raw_workspace or "data/workspace"))
    return p.resolve() if p.is_absolute() else (Path.cwd() / p).resolve()


def _safe_name(name: str) -> str:
    n = re.sub(r"[^\w.\-]+", "_", str(name or "").strip())
    return n[:180] if n else f"file_{uuid.uuid4().hex[:8]}"


def _incoming_dir(workspace: Path, platform: str) -> Path:
    day = datetime.now().strftime("%Y%m%d")
    d = workspace / "incoming" / platform / day
    d.mkdir(parents=True, exist_ok=True)
    return d


def _delivery_log(workspace: Path, platform: str, chat_id: Any, status: str, detail: str = "") -> None:
    try:
        ts = datetime.now().isoformat()
        line = f"{ts}\t{platform}\t{chat_id}\t{status}\t{str(detail or '').strip()[:500]}\n"
        targets = []
        with contextlib.suppress(Exception):
            targets.append(((workspace / "logs").resolve() / "messenger_delivery.log"))
        with contextlib.suppress(Exception):
            repo_root = Path(__file__).resolve().parents[2]
            targets.append((repo_root / "data" / "logs" / "messenger_delivery.log").resolve())
        written: set[str] = set()
        for p in targets:
            key = str(p).lower()
            if key in written:
                continue
            written.add(key)
            with contextlib.suppress(Exception):
                p.parent.mkdir(parents=True, exist_ok=True)
                with p.open("a", encoding="utf-8") as fh:
                    fh.write(line)
    except Exception:
        pass


def _extract_paths_for_send(reply: str, workspace: Optional[Path] = None, max_files: int = 3) -> List[Path]:
    text = str(reply or "")
    lower = text.lower()
    send_triggers = [
        "gespeichert unter",
        "saved to",
        "hier ist die datei",
        "anbei",
        "im anhang",
        "download",
        "datei findest du hier",
        "ich schicke dir die datei",
        "ich sende dir die datei",
        "wird gesendet",
        "wird geschickt",
        "sending file",
    ]
    # Avoid auto-sending files for generic script creation confirmations.
    if not any(t in lower for t in send_triggers):
        return []
    candidates: List[str] = []
    candidates.extend(re.findall(r"`([^`]+)`", text))
    candidates.extend(re.findall(r"([A-Za-z]:\\[^\s<>\"']+)", text))
    candidates.extend(re.findall(r"\"([A-Za-z]:\\[^\"]+)\"", text))
    candidates.extend(re.findall(r"\b([A-Za-z0-9._-]+\.[A-Za-z0-9]{2,8})\b", text))
    out: List[Path] = []
    seen: set[str] = set()
    blocked_ext = {".py", ".ps1", ".bat", ".cmd", ".sh", ".ogg", ".oga", ".mp3", ".wav", ".m4a", ".opus"}

    def _resolve_candidate(raw: str) -> Optional[Path]:
        p = Path(str(raw or "").strip())
        if not str(p):
            return None
        if p.is_absolute():
            if p.exists() and p.is_file():
                return p.resolve()
            return None
        # Resolve relative names against workspace by newest match first.
        ws = workspace.resolve() if isinstance(workspace, Path) else None
        if ws and ws.exists():
            best: Optional[Path] = None
            best_mtime = -1.0
            for fp in ws.rglob("*"):
                try:
                    if not fp.is_file():
                        continue
                    if fp.name.lower() != p.name.lower():
                        continue
                    mt = float(fp.stat().st_mtime)
                    if mt > best_mtime:
                        best_mtime = mt
                        best = fp.resolve()
                except Exception:
                    continue
            if best is not None:
                return best
        # Fallback to current working directory only if workspace match was not found.
        if p.exists() and p.is_file():
            return p.resolve()
        return None

    for c in candidates:
        p = _resolve_candidate(c)
        if p is None or not p.exists() or not p.is_file():
            continue
        if p.suffix.lower() in blocked_ext:
            # Scripts only by explicit command, not auto by path mention.
            continue
        key = str(p.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p.resolve())
        if len(out) >= max_files:
            break
    return out


_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
_IMAGE_MAX_BYTES = 3 * 1024 * 1024  # 3 MB Rohdaten (base64 ~+33%)


def _encode_image(path: str) -> Optional[Dict[str, Any]]:
    """Kodiert ein Bild als base64 für LLM-Vision-Input.
    Returns None wenn kein Bild-Format.
    Returns {"too_large": True, "filename": ..., "size_mb": ...} wenn Bild zu groß.
    Returns {"media_type": ..., "data": ..., "filename": ...} bei Erfolg."""
    import base64, mimetypes
    p = Path(path)
    if p.suffix.lower() not in _IMAGE_EXTS:
        return None
    try:
        raw = p.read_bytes()
        if len(raw) > _IMAGE_MAX_BYTES:
            return {"too_large": True, "filename": p.name, "size_mb": round(len(raw) / 1024 / 1024, 1)}
        mime = mimetypes.guess_type(str(p))[0] or "image/jpeg"
        return {"media_type": mime, "data": base64.b64encode(raw).decode(), "filename": p.name}
    except Exception:
        return None


def _append_attachment_context(text: str, attachments: List[str]) -> str:
    base = str(text or "").strip()
    if not attachments:
        return base
    if not base:
        base = "Dateianhang empfangen."
    lines = [base, "", "Anhaenge:"]
    for p in attachments:
        lines.append(f"- {p}")
    return "\n".join(lines)


_DOC_TEXT_MAX_CHARS = 8000  # Max Zeichen für injizierte Dateiinhalte


def _extract_document_content(path: str) -> Optional[str]:
    """Extrahiert Textinhalt aus Dokumenten für LLM-Kontext.
    Unterstützt: PDF, TXT, CSV, DOCX, XLSX.
    Gibt None zurück wenn Format nicht unterstützt oder Fehler."""
    p = Path(path)
    ext = p.suffix.lower()

    try:
        # --- Plaintext ---
        if ext in {".txt", ".md", ".log", ".py", ".js", ".json", ".xml", ".html", ".css"}:
            content = p.read_text(encoding="utf-8", errors="replace").strip()
            if content:
                return content[:_DOC_TEXT_MAX_CHARS]

        # --- CSV ---
        elif ext == ".csv":
            import csv as _csv
            rows = []
            with p.open(encoding="utf-8", errors="replace", newline="") as f:
                reader = _csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 50:
                        rows.append(f"... (nur erste 50 Zeilen gezeigt)")
                        break
                    rows.append(", ".join(row))
            return "\n".join(rows)[:_DOC_TEXT_MAX_CHARS] if rows else None

        # --- PDF ---
        elif ext == ".pdf":
            # Versuch 1: pypdf
            try:
                from pypdf import PdfReader
                reader = PdfReader(str(p))
                pages_text = []
                for pg in reader.pages:
                    t = pg.extract_text()
                    if t:
                        pages_text.append(t.strip())
                content = "\n\n".join(pages_text).strip()
                if content:
                    return content[:_DOC_TEXT_MAX_CHARS]
            except ImportError:
                pass
            # Versuch 2: pdfplumber
            try:
                import pdfplumber
                with pdfplumber.open(str(p)) as pdf:
                    pages_text = [pg.extract_text() or "" for pg in pdf.pages]
                content = "\n\n".join(pages_text).strip()
                if content:
                    return content[:_DOC_TEXT_MAX_CHARS]
            except ImportError:
                pass

        # --- DOCX ---
        elif ext == ".docx":
            try:
                import docx as _docx
                doc = _docx.Document(str(p))
                content = "\n".join(para.text for para in doc.paragraphs if para.text.strip())
                if content:
                    return content[:_DOC_TEXT_MAX_CHARS]
            except ImportError:
                pass

        # --- XLSX ---
        elif ext in {".xlsx", ".xls"}:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                rows_out = []
                for sheet in wb.sheetnames[:3]:  # Max 3 Sheets
                    ws = wb[sheet]
                    rows_out.append(f"[Sheet: {sheet}]")
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= 50:
                            rows_out.append("... (nur erste 50 Zeilen)")
                            break
                        if any(c is not None for c in row):
                            rows_out.append("\t".join(str(c) if c is not None else "" for c in row))
                content = "\n".join(rows_out)
                if content.strip():
                    return content[:_DOC_TEXT_MAX_CHARS]
            except ImportError:
                pass

    except Exception:
        pass

    return None


class TelegramWorker(threading.Thread):
    def __init__(
        self,
        token: str,
        submit_fn: Callable,  # Geändert von reply_fn zu submit_fn (Queue)
        log_fn: LogFn,
        workspace: str,
        audio_service: Optional[AudioService] = None,
        tts_cfg: Optional[Dict[str, Any]] = None,
        reply_timeout_sec: int = 600,
    ):
        super().__init__(daemon=True, name="K.AI Agent-telegram-worker")
        self.token = token
        self.submit_fn = submit_fn
        self.log_fn = log_fn
        self.stop_event = threading.Event()
        self.base = f"https://api.telegram.org/bot{token}"
        self.workspace = _workspace_root(workspace)
        self.audio_service = audio_service
        self.tts_cfg = tts_cfg or {}
        self.tts_enabled = bool(self.tts_cfg.get("enabled", False))
        self.tts_voice = str(self.tts_cfg.get("voice", "de-DE-ConradNeural"))
        self.reply_timeout_sec = max(10, min(3600, int(reply_timeout_sec or 600)))
        self.offset = self._load_offset()
        self._last_send_error = ""
        self._stats_lock = threading.Lock()
        self._pending_outbox_count = 0
        self._failed_update_count = 0
        self._last_delivery_state = ""

    def _offset_path(self) -> Path:
        return self.workspace / ".tg_offset"

    def _load_offset(self) -> int:
        try:
            p = self._offset_path()
            if p.exists():
                return int(p.read_text(encoding="utf-8").strip())
        except Exception:
            pass
        return 0

    def _save_offset(self) -> None:
        try:
            p = self._offset_path()
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(str(self.offset), encoding="utf-8")
        except Exception as exc:
            self.log_fn(f"Telegram Offset-Speichern fehlgeschlagen ({self.offset}): {exc}")

    def _bootstrap_offset_if_missing(self, client: httpx.Client) -> None:
        """
        Verhindert Replay alter Telegram-Nachrichten bei Neustart.
        Wenn keine gespeicherte Offset-Datei vorhanden ist (offset<=0),
        wird der bestehende Bot-Backlog einmalig verworfen.
        """
        if int(self.offset or 0) > 0:
            return
        latest_update_id: Optional[int] = None
        probe_offset = 0
        scanned = 0
        try:
            for _ in range(10):
                res = client.get(
                    f"{self.base}/getUpdates",
                    params={
                        "offset": probe_offset,
                        "timeout": 0,
                        "limit": 100,
                        "allowed_updates": ["message"],
                    },
                    timeout=20,
                )
                res.raise_for_status()
                body = res.json() if isinstance(res.json(), dict) else {}
                if not bool(body.get("ok", False)):
                    break
                updates = body.get("result", []) if isinstance(body.get("result", []), list) else []
                if not updates:
                    break
                scanned += len(updates)
                for upd in updates:
                    try:
                        uid = int(upd.get("update_id", 0))
                    except Exception:
                        uid = 0
                    if uid > 0 and (latest_update_id is None or uid > latest_update_id):
                        latest_update_id = uid
                if isinstance(latest_update_id, int):
                    probe_offset = latest_update_id + 1
                if len(updates) < 100:
                    break
        except Exception as exc:
            self.log_fn(f"Telegram Backlog-Bootstrap fehlgeschlagen: {exc}")
            return
        if isinstance(latest_update_id, int):
            self.offset = latest_update_id + 1
            self._save_offset()
            self.log_fn(
                f"Telegram Backlog verworfen (erstes Starten ohne Offset): "
                f"new_offset={self.offset}, updates_seen={scanned}"
            )
            _delivery_log(
                self.workspace,
                "telegram",
                "-",
                "bootstrap_flush",
                f"new_offset={self.offset} updates_seen={scanned}",
            )


    def stop(self) -> None:
        self.stop_event.set()

    def _safe_post_json(self, client: httpx.Client, path: str, payload: Dict[str, Any]) -> bool:
        """Sendet JSON an Telegram API mit Retry-Logik."""
        import time
        max_retries = 3
        retry_delay = 1
        self._last_send_error = ""
        
        for attempt in range(max_retries):
            try:
                resp = client.post(f"{self.base}/{path}", json=payload, timeout=20)
                
                if resp.status_code == 200:
                    self._last_send_error = ""
                    return True  # Erfolgreich
                elif resp.status_code == 429:
                    # Rate-Limiting
                    retry_after = resp.json().get("parameters", {}).get("retry_after", 5)
                    self.log_fn(f"Telegram Rate-Limit, warte {retry_after}s")
                    self._last_send_error = f"rate_limit:{retry_after}"
                    time.sleep(retry_after)
                    continue
                elif attempt < max_retries - 1:
                    self.log_fn(f"Telegram {path} Versuch {attempt + 1} fehlgeschlagen (Status {resp.status_code}), retry...")
                    body_short = ""
                    with contextlib.suppress(Exception):
                        body_short = str(resp.text or "")[:200]
                    self._last_send_error = f"http_{resp.status_code}:{body_short}"
                    time.sleep(retry_delay * (attempt + 1))
                else:
                    self.log_fn(f"Telegram {path} nach {max_retries} Versuchen fehlgeschlagen (Status {resp.status_code})")
                    body_short = ""
                    with contextlib.suppress(Exception):
                        body_short = str(resp.text or "")[:200]
                    self._last_send_error = f"http_{resp.status_code}:{body_short}"
            except Exception as e:
                if attempt < max_retries - 1:
                    self.log_fn(f"Telegram {path} Versuch {attempt + 1} Exception: {e}, retry...")
                    self._last_send_error = f"exception:{e}"
                    time.sleep(retry_delay * (attempt + 1))
                else:
                    self.log_fn(f"Telegram {path} nach {max_retries} Versuchen Exception: {e}")
                    self._last_send_error = f"exception:{e}"
        return False

    def get_stats(self) -> Dict[str, Any]:
        with self._stats_lock:
            return {
                "pending_outbox": int(self._pending_outbox_count),
                "failed_updates": int(self._failed_update_count),
                "last_delivery_state": str(self._last_delivery_state or ""),
            }

    def _send_text_chunks(self, client: httpx.Client, chat_id: Any, text: str) -> bool:
        if not text:
            return False
            
        chunk_size = 4000
        sent_any = False
        all_ok = True
        for i in range(0, len(text), chunk_size):
            ok = self._safe_post_json(client, "sendMessage", {
                "chat_id": chat_id, 
                "text": text[i : i + chunk_size]
            })
            sent_any = sent_any or bool(ok)
            if not ok:
                all_ok = False
                break
        return sent_any and all_ok

    def _send_document(self, client: httpx.Client, chat_id: Any, path: Path) -> None:
        try:
            with path.open("rb") as fh:
                files = {"document": (path.name, fh, "application/octet-stream")}
                data = {"chat_id": str(chat_id)}
                client.post(f"{self.base}/sendDocument", data=data, files=files, timeout=60)
        except Exception as exc:
            self.log_fn(f"Telegram sendDocument Fehler ({path}): {exc}")

    def _send_voice(self, client: httpx.Client, chat_id: Any, path: Path) -> None:
        try:
            # Telegram expects OGG/Opus for voice messages generally, but accepts MP3/others as audio
            # Using sendVoice for voice messages (renders as waveform)
            with path.open("rb") as fh:
                files = {"voice": (path.name, fh, "audio/ogg")}
                data = {"chat_id": str(chat_id)}
                client.post(f"{self.base}/sendVoice", data=data, files=files, timeout=60)
        except Exception as exc:
            self.log_fn(f"Telegram sendVoice Fehler ({path}): {exc}")

    def _get_file_bytes(self, client: httpx.Client, file_id: str) -> tuple[bytes, str]:
        res = client.get(f"{self.base}/getFile", params={"file_id": file_id}, timeout=25)
        res.raise_for_status()
        obj = res.json().get("result", {}) if isinstance(res.json(), dict) else {}
        file_path = str(obj.get("file_path", "") or "")
        if not file_path:
            raise RuntimeError("telegram_file_path_missing")
        file_url = f"https://api.telegram.org/file/bot{self.token}/{file_path}"
        bin_res = client.get(file_url, timeout=60)
        bin_res.raise_for_status()
        return bin_res.content, file_path

    def _download_attachments(self, client: httpx.Client, msg: Dict[str, Any]) -> List[str]:
        out_paths: List[str] = []
        entries: List[tuple[str, str]] = []
        if isinstance(msg.get("document"), dict):
            d = msg["document"]
            entries.append((str(d.get("file_id", "")), str(d.get("file_name", "") or "")))
        if isinstance(msg.get("audio"), dict):
            d = msg["audio"]
            entries.append((str(d.get("file_id", "")), str(d.get("file_name", "") or "audio.mp3")))
        if isinstance(msg.get("video"), dict):
            d = msg["video"]
            entries.append((str(d.get("file_id", "")), str(d.get("file_name", "") or "video.mp4")))
        if isinstance(msg.get("voice"), dict):
            d = msg["voice"]
            entries.append((str(d.get("file_id", "")), "voice.ogg"))
        if isinstance(msg.get("animation"), dict):
            d = msg["animation"]
            entries.append((str(d.get("file_id", "")), str(d.get("file_name", "") or "animation.mp4")))
        photos = msg.get("photo")
        if isinstance(photos, list) and photos:
            p = photos[-1]
            if isinstance(p, dict):
                entries.append((str(p.get("file_id", "")), "photo.jpg"))

        if not entries:
            return out_paths

        target_dir = _incoming_dir(self.workspace, "telegram")
        for file_id, raw_name in entries:
            if not file_id:
                continue
            try:
                data, remote_path = self._get_file_bytes(client, file_id)
                ext = Path(remote_path).suffix or Path(raw_name).suffix
                stem = Path(raw_name).stem or Path(remote_path).stem or f"tg_{uuid.uuid4().hex[:8]}"
                name = f"{_safe_name(stem)}{ext}"
                target = (target_dir / name).resolve()
                target.write_bytes(data)
                out_paths.append(str(target))
            except Exception as exc:
                self.log_fn(f"Telegram download attachment Fehler: {exc}")
        return out_paths

    def run(self) -> None:
        self.log_fn("Telegram Worker gestartet.")
        failed_updates: Dict[int, int] = {}
        pending_outbox: List[Dict[str, Any]] = []
        queued_update_ids: set[int] = set()
        with httpx.Client(timeout=30, trust_env=False) as client:
            # Nur beim allerersten Start ohne bestehende .tg_offset:
            # alte Updates nicht erneut beantworten.
            self._bootstrap_offset_if_missing(client)
            while not self.stop_event.is_set():
                try:
                    # Best-effort resend queue: verhindert, dass ein einzelner Sendefehler
                    # den gesamten Update-Stream blockiert.
                    if pending_outbox:
                        now_ts = time.time()
                        remaining: List[Dict[str, Any]] = []
                        for item in pending_outbox:
                            try:
                                next_try = float(item.get("next_try", 0.0) or 0.0)
                            except Exception:
                                next_try = 0.0
                            if next_try > now_ts:
                                remaining.append(item)
                                continue
                            chat_id = item.get("chat_id")
                            text = str(item.get("text", "") or "")
                            upd_ref = int(item.get("update_id", 0) or 0)
                            if not chat_id or not text:
                                if upd_ref:
                                    queued_update_ids.discard(upd_ref)
                                continue
                            ok_retry = self._send_text_chunks(client, chat_id, text)
                            if ok_retry:
                                _delivery_log(self.workspace, "telegram", chat_id, "resend_ok", text[:200])
                                if upd_ref:
                                    queued_update_ids.discard(upd_ref)
                                continue
                            attempts = int(item.get("attempts", 0) or 0) + 1
                            err_hint = str(getattr(self, "_last_send_error", "") or "").strip()
                            if attempts >= 8:
                                _delivery_log(
                                    self.workspace,
                                    "telegram",
                                    chat_id,
                                    "resend_dropped",
                                    f"{text[:200]} | retries={attempts} | err={err_hint[:220]}",
                                )
                                self.log_fn(
                                    f"Telegram resend dropped (update_id={upd_ref}, retries={attempts}, err={err_hint or 'unknown'})"
                                )
                                if upd_ref:
                                    queued_update_ids.discard(upd_ref)
                                continue
                            backoff = float(min(120, 2 ** min(6, attempts)))
                            item["attempts"] = attempts
                            item["next_try"] = time.time() + backoff
                            remaining.append(item)
                            _delivery_log(
                                self.workspace,
                                "telegram",
                                chat_id,
                                "resend_failed",
                                f"{text[:200]} | retries={attempts} | next={int(backoff)}s | err={err_hint[:220]}",
                            )
                        pending_outbox = remaining
                    with self._stats_lock:
                        self._pending_outbox_count = len(pending_outbox)
                        self._failed_update_count = len(failed_updates)

                    res = client.get(
                        f"{self.base}/getUpdates",
                        params={"offset": self.offset, "timeout": 20, "allowed_updates": ["message"]},
                    )
                    res.raise_for_status()
                    data = res.json()
                    updates = data.get("result", []) if data.get("ok") else []
                    for upd in updates:
                        upd_id = int(upd.get("update_id", 0))
                        next_offset = upd_id + 1
                        ack_update = True
                        try:
                            msg = upd.get("message") or {}
                            from_user = msg.get("from") or {}
                            if bool(from_user.get("is_bot")):
                                if next_offset > self.offset:
                                    self.offset = next_offset
                                    self._save_offset()
                                continue
                            chat = msg.get("chat") or {}
                            chat_id = chat.get("id")
                            if not chat_id:
                                if next_offset > self.offset:
                                    self.offset = next_offset
                                    self._save_offset()
                                continue
                            user = from_user.get("username") or from_user.get("first_name") or "unknown"
                            text = str(msg.get("text") or msg.get("caption") or "").strip()
                            attachments = self._download_attachments(client, msg)

                            # SST Logic
                            transcriptions = []
                            user_sent_voice = False
                            if self.audio_service:
                                for att in attachments:
                                    p = Path(att)
                                    suffix = p.suffix.lower()
                                    self.log_fn(f"SST: checking attachment: {p}, suffix: {suffix}")
                                    if suffix in [".ogg", ".oga", ".mp3", ".wav", ".m4a", ".opus"]:
                                        user_sent_voice = True
                                        if not p.exists():
                                            self.log_fn(f"SST: file does not exist: {p}")
                                            continue
                                        self._safe_post_json(client, "sendChatAction", {"chat_id": chat_id, "action": "upload_voice"})
                                        tr = self.audio_service.transcribe(p)
                                        self.log_fn(f"SST: result for {p.name}: {tr[:100] if tr else 'empty'}")
                                        if tr and not tr.startswith("[SST Error"):
                                            # Use raw transcription without prefix to avoid echo
                                            transcriptions.append(tr)
                                        elif tr.startswith("[SST Error"):
                                            self.log_fn(f"SST Error: {tr}")
                        
                            audio_suffixes = {".ogg", ".oga", ".mp3", ".wav", ".m4a", ".opus"}
                            # Audio-Dateien immer aus Anhangsliste entfernen – entweder transkribiert oder ignoriert
                            non_audio_attachments = [a for a in attachments if Path(a).suffix.lower() not in audio_suffixes]
                            if transcriptions:
                                # Replace or append transcription to text
                                if text.strip():
                                    text = (text + "\n\n" + "\n".join(transcriptions)).strip()
                                else:
                                    text = "\n".join(transcriptions).strip()
                            elif user_sent_voice and not text.strip():
                                # Transkription fehlgeschlagen und kein Text vorhanden → abbrechen statt Pfad senden
                                self.log_fn("SST: Transkription fehlgeschlagen, keine Antwort gesendet")
                                self._safe_post_json(client, "sendMessage", {"chat_id": chat_id, "text": "⚠️ Spracherkennung fehlgeschlagen. Bitte erneut versuchen oder als Text schreiben."})
                                if next_offset > self.offset:
                                    self.offset = next_offset
                                    self._save_offset()
                                continue

                            # Bilder für Vision-Analyse enkodieren (max 3 MB, nur bekannte Formate)
                            # Dokumente werden per Text-Extraktion in den Kontext injiziert
                            _tg_image_blobs: List[Dict[str, Any]] = []
                            _tg_non_image_paths: List[str] = []
                            _tg_too_large_hints: List[str] = []
                            _tg_doc_contents: List[str] = []
                            for _ap in non_audio_attachments:
                                _blob = _encode_image(_ap)
                                if _blob is None:
                                    # Kein Bild — versuche Dokumenteninhalt zu extrahieren
                                    _doc_text = _extract_document_content(_ap)
                                    if _doc_text:
                                        _fname = Path(_ap).name
                                        _tg_doc_contents.append(f"📄 Inhalt von '{_fname}':\n{_doc_text}")
                                    else:
                                        _tg_non_image_paths.append(_ap)
                                elif _blob.get("too_large"):
                                    _tg_too_large_hints.append(f"⚠️ Bild '{_blob['filename']}' ({_blob['size_mb']} MB) ist zu groß für Bildanalyse (max 3 MB) und wurde als Datei gespeichert.")
                                    _tg_non_image_paths.append(_ap)
                                else:
                                    _tg_image_blobs.append(_blob)

                            effective_text = _append_attachment_context(text, _tg_non_image_paths)
                            if _tg_doc_contents:
                                _doc_block = "\n\n".join(_tg_doc_contents)
                                effective_text = (effective_text + "\n\n" + _doc_block).strip() if effective_text else _doc_block
                            if _tg_too_large_hints:
                                effective_text = (effective_text + "\n\n" + "\n".join(_tg_too_large_hints)).strip()
                            if not effective_text and not _tg_image_blobs:
                                if next_offset > self.offset:
                                    self.offset = next_offset
                                    self._save_offset()
                                continue
                            if not effective_text:
                                effective_text = "Bild erhalten."

                            # --- Asynchrone Queue Einreichung mit Live-Updates ---
                            status_msg_id = None
                            
                            def _send_status_update(status: str):
                                """Sendet einen Status-Update an Telegram"""
                                nonlocal status_msg_id
                                try:
                                    with httpx.Client(timeout=10, trust_env=False) as c:
                                        if status_msg_id:
                                            # Bearbeite existierende Nachricht
                                            c.post(
                                                f"{self.base}/editMessageText",
                                                json={
                                                    "chat_id": chat_id,
                                                    "message_id": status_msg_id,
                                                    "text": status
                                                }
                                            )
                                        else:
                                            # Erstelle neue Status-Nachricht
                                            resp = c.post(
                                                f"{self.base}/sendMessage",
                                                json={"chat_id": chat_id, "text": status}
                                            )
                                            if resp.status_code == 200:
                                                data = resp.json()
                                                if data.get("ok"):
                                                    status_msg_id = data.get("result", {}).get("message_id")
                                except Exception:
                                    pass
                            
                            # Sende initiale Status-Nachricht
                            _send_status_update("⏳ Verarbeite deine Anfrage...")
                            
                            def _on_finish_tg(res: Dict[str, Any]):
                                reply_text = _ensure_reply_text(res.get("reply", ""))
                                steps = res.get("steps", [])
                                
                                with httpx.Client(timeout=30, trust_env=False) as c:
                                    # Bei Sprachnachricht: TTS falls aktiv, sonst Text
                                    if user_sent_voice and self.audio_service and self.tts_enabled:
                                        # Status-Nachricht (⏳) löschen bevor Audio kommt, da Audio neuer Typ ist
                                        if status_msg_id:
                                            with contextlib.suppress(Exception):
                                                c.post(
                                                    f"{self.base}/deleteMessage",
                                                    json={"chat_id": chat_id, "message_id": status_msg_id}
                                                )
                                        try:
                                            out_path = self.workspace / "temp" / f"tts_{uuid.uuid4().hex[:8]}.mp3"
                                            out_path.parent.mkdir(parents=True, exist_ok=True)
                                            c.post(f"{self.base}/sendChatAction", json={"chat_id": chat_id, "action": "record_voice"})
                                            try:
                                                import asyncio
                                                loop = asyncio.new_event_loop()
                                                audio_file = loop.run_until_complete(self.audio_service.synthesize(reply_text[:4000], out_path, self.tts_voice))
                                                loop.close()
                                                if audio_file and audio_file.exists():
                                                    self._send_voice(c, chat_id, audio_file)
                                                    with contextlib.suppress(Exception): audio_file.unlink()
                                            except Exception as e:
                                                self.log_fn(f"Telegram TTS Error: {e}")
                                        except Exception: pass
                                    else:
                                        # Finale Antwort: Edit nur wenn keine Schritt-Nachrichten gesendet wurden,
                                        # sonst würde die Antwort über den Schritt-Meldungen erscheinen
                                        sent_via_edit = False
                                        has_steps = bool(steps)

                                        if status_msg_id and not has_steps and len(reply_text) < 4000:
                                            resp = c.post(
                                                f"{self.base}/editMessageText",
                                                json={
                                                    "chat_id": chat_id,
                                                    "message_id": status_msg_id,
                                                    "text": reply_text
                                                }
                                            )
                                            tg_data = resp.json()
                                            if resp.status_code == 200 and tg_data.get("ok"):
                                                sent_via_edit = True
                                                try:
                                                    from app.tool_engine import tool_store
                                                    tool_store.log("telegram_message_sent", f"Nachricht editiert: msg_id={status_msg_id}")
                                                except Exception:
                                                    pass
                                        
                                        if not sent_via_edit:
                                            # Fallback oder wenn Nachricht zu lang: Status löschen und neu senden
                                            if status_msg_id:
                                                with contextlib.suppress(Exception):
                                                    c.post(f"{self.base}/deleteMessage", json={"chat_id": chat_id, "message_id": status_msg_id})
                                            self._send_text_chunks(c, chat_id, reply_text)
                                    
                                    # Auto-Send nur wenn der Agent keine Datei via Tool gesendet hat
                                    agent_sent_file = any(
                                        isinstance(step, dict)
                                        and step.get("kind") == "send_messenger_file"
                                        and step.get("ok")
                                        for step in steps
                                    )
                                    if not agent_sent_file:
                                        for p in _extract_paths_for_send(reply_text, workspace=self.workspace):
                                            self._send_document(c, chat_id, p)

                            # Request an Queue übergeben (Kein Blockieren!)
                            _tg_submit_kwargs: Dict[str, Any] = dict(
                                result_callback=_on_finish_tg,
                                source="messenger:telegram",
                                dialog_key=f"telegram:{chat_id}",
                                platform="telegram",
                                chat_id=chat_id,
                                user=user,
                            )
                            if _tg_image_blobs:
                                _tg_submit_kwargs["images"] = _tg_image_blobs
                            self.submit_fn(effective_text, **_tg_submit_kwargs)
                            # Rückmeldung an User sofort senden
                            # self._safe_post_json(client, "sendMessage", {"chat_id": chat_id, "text": "Ich habe deine Anfrage erhalten und bearbeite sie..."})
                            ack_update = True

                        except Exception as upd_exc:
                            retry_count = int(failed_updates.get(upd_id, 0) or 0) + 1
                            failed_updates[upd_id] = retry_count
                            if retry_count >= 3:
                                ack_update = True
                                self.log_fn(
                                    f"Telegram update processing dropped after {retry_count} tries (update_id={upd_id}): {upd_exc}"
                                )
                                with self._stats_lock:
                                    self._last_delivery_state = f"update_dropped:update={upd_id}"
                            else:
                                ack_update = False
                                self.log_fn(
                                    f"Telegram update processing error (update_id={upd_id}, retry={retry_count}): {upd_exc}"
                                )
                                with self._stats_lock:
                                    self._last_delivery_state = f"update_error:update={upd_id}"
                        finally:
                            if ack_update and next_offset > self.offset:
                                self.offset = next_offset
                                self._save_offset()
                                failed_updates.pop(upd_id, None)
                            with self._stats_lock:
                                self._pending_outbox_count = len(pending_outbox)
                                self._failed_update_count = len(failed_updates)
                except Exception as exc:
                    self.log_fn(f"Telegram Worker Fehler: {exc}")
                    time.sleep(2)
        self.log_fn("Telegram Worker gestoppt.")


class DiscordRestWorker(threading.Thread):
    def __init__(
        self,
        token: str,
        channel_id: str,
        submit_fn: Callable, # Geändert
        log_fn: LogFn,
        workspace: str,
        audio_service: Optional[AudioService] = None,
        tts_cfg: Optional[Dict[str, Any]] = None,
        reply_timeout_sec: int = 600,
    ):
        super().__init__(daemon=True, name="K.AI Agent-discord-worker")
        self.token = token
        self.channel_id = channel_id
        self.submit_fn = submit_fn
        self.log_fn = log_fn
        self.stop_event = threading.Event()
        self.last_seen: Optional[int] = None
        self.base = "https://discord.com/api/v10"
        self.workspace = _workspace_root(workspace)
        # REST Worker does not support SST/TTS fully yet (polling limitation), passing args for compatibility
        self.audio_service = audio_service
        self.tts_cfg = tts_cfg or {}
        self.reply_timeout_sec = max(10, min(3600, int(reply_timeout_sec or 600)))


    def stop(self) -> None:
        self.stop_event.set()

    def _headers_json(self) -> Dict[str, str]:
        return {"Authorization": f"Bot {self.token}", "Content-Type": "application/json"}

    def _headers_auth(self) -> Dict[str, str]:
        return {"Authorization": f"Bot {self.token}"}

    def _download_attachments(self, client: httpx.Client, msg: Dict[str, Any]) -> List[str]:
        out_paths: List[str] = []
        atts = msg.get("attachments") or []
        if not isinstance(atts, list) or not atts:
            return out_paths
        target_dir = _incoming_dir(self.workspace, "discord")
        for a in atts:
            if not isinstance(a, dict):
                continue
            url = str(a.get("url", "") or "")
            if not url:
                continue
            name = _safe_name(str(a.get("filename", "") or f"dc_{uuid.uuid4().hex[:8]}"))
            target = (target_dir / name).resolve()
            try:
                r = client.get(url, timeout=60)
                r.raise_for_status()
                target.write_bytes(r.content)
                out_paths.append(str(target))
            except Exception as exc:
                self.log_fn(f"Discord attachment download Fehler: {exc}")
        return out_paths

    def _send_text(self, client: httpx.Client, content: str) -> None:
        msg = str(content or "")
        if not msg:
            return
        chunk_size = 1900
        for i in range(0, len(msg), chunk_size):
            client.post(
                f"{self.base}/channels/{self.channel_id}/messages",
                headers=self._headers_json(),
                json={"content": msg[i : i + chunk_size]},
                timeout=20,
            )

    def _send_file(self, client: httpx.Client, path: Path) -> None:
        try:
            with path.open("rb") as fh:
                files = {"files[0]": (path.name, fh, "application/octet-stream")}
                payload_json = json.dumps({"content": ""})
                client.post(
                    f"{self.base}/channels/{self.channel_id}/messages",
                    headers=self._headers_auth(),
                    data={"payload_json": payload_json},
                    files=files,
                    timeout=60,
                )
        except Exception as exc:
            self.log_fn(f"Discord send file Fehler ({path}): {exc}")

    def run(self) -> None:
        self.log_fn(f"Discord Worker gestartet (channel_id={self.channel_id}).")
        with httpx.Client(timeout=20, trust_env=False) as client:
            try:
                init_res = client.get(
                    f"{self.base}/channels/{self.channel_id}/messages",
                    params={"limit": 1},
                    headers=self._headers_json(),
                )
                init_res.raise_for_status()
                items = init_res.json()
                if items:
                    self.last_seen = int(items[0].get("id", "0"))
            except Exception as exc:
                self.log_fn(f"Discord Initialisierung fehlgeschlagen: {exc}")

            while not self.stop_event.is_set():
                try:
                    res = client.get(
                        f"{self.base}/channels/{self.channel_id}/messages",
                        params={"limit": 20},
                        headers=self._headers_json(),
                    )
                    res.raise_for_status()
                    items = res.json() or []
                    items_sorted = sorted(items, key=lambda x: int(x.get("id", "0")))
                    for msg in items_sorted:
                        mid = int(msg.get("id", "0"))
                        if self.last_seen is not None and mid <= self.last_seen:
                            continue
                        self.last_seen = mid
                        author = msg.get("author") or {}
                        if author.get("bot"):
                            continue
                        text = str(msg.get("content") or "").strip()
                        attachments = self._download_attachments(client, msg)
                        _dc_non_doc: List[str] = []
                        _dc_doc_contents: List[str] = []
                        for _ap in attachments:
                            _doc_text = _extract_document_content(_ap)
                            if _doc_text:
                                _dc_doc_contents.append(f"📄 Inhalt von '{Path(_ap).name}':\n{_doc_text}")
                            else:
                                _dc_non_doc.append(_ap)
                        effective_text = _append_attachment_context(text, _dc_non_doc)
                        if _dc_doc_contents:
                            _dc_doc_block = "\n\n".join(_dc_doc_contents)
                            effective_text = (effective_text + "\n\n" + _dc_doc_block).strip() if effective_text else _dc_doc_block
                        if not effective_text:
                            continue
                        # Typing-Indikator: alle 8s erneuern (Discord-Limit: ~10s)
                        _typing_stop = threading.Event()
                        def _keep_typing_dc():
                            while not _typing_stop.is_set():
                                with contextlib.suppress(Exception):
                                    client.post(
                                        f"{self.base}/channels/{self.channel_id}/typing",
                                        headers=self._headers_json(),
                                        timeout=8,
                                    )
                                _typing_stop.wait(8)
                        try:
                            start_thread_with_trace(_keep_typing_dc, name=f"K.AI-discord-typing")
                        except Exception:
                            _typing_thread = threading.Thread(target=_keep_typing_dc, daemon=True)
                            _typing_thread.start()
                        user = author.get("username") or "unknown"
                        # --- Asynchrone Queue Einreichung (Discord Rest) ---
                        def _on_finish_dc_rest(res: Dict[str, Any]):
                            reply_text = _ensure_reply_text(res.get("reply", ""))
                            steps_dc = res.get("steps", [])
                            with httpx.Client(timeout=20, trust_env=False) as c:
                                self._send_text(c, reply_text)
                                agent_sent_file_dc = any(
                                    isinstance(s, dict) and s.get("kind") == "send_messenger_file" and s.get("ok")
                                    for s in steps_dc
                                )
                                if not agent_sent_file_dc:
                                    for p in _extract_paths_for_send(reply_text, workspace=self.workspace):
                                        self._send_file(c, p)

                        # In Queue schieben
                        self.submit_fn(
                            effective_text,
                            result_callback=_on_finish_dc_rest,
                            source="messenger:discord-rest",
                            dialog_key=f"dc:{self.channel_id}",
                            platform="discord",
                            channel_id=self.channel_id,
                            user=user
                        )
                        # Sofortbestätigung (Optional, Discord-Rest ist oft für Monitoring)
                        # self._send_text(client, "_Anfrage wird verarbeitet..._")
                except Exception as exc:
                    self.log_fn(f"Discord Worker Fehler: {exc}")
                    time.sleep(2)
                time.sleep(1.5)
        self.log_fn("Discord Worker gestoppt.")




class DiscordGatewayWorker(threading.Thread):
    def __init__(
        self,
        token: str,
        channel_id: str,
        guild_id: str,
        submit_fn: Callable, # Geändert
        log_fn: LogFn,
        workspace: str,
        presence_cfg: Optional[Dict[str, Any]] = None,
        audio_service: Optional[AudioService] = None,
        tts_cfg: Optional[Dict[str, Any]] = None,
        reply_timeout_sec: int = 600,
    ):
        super().__init__(daemon=True, name="K.AI Agent-discord-gateway-worker")
        self.token = token
        self.channel_id = channel_id
        self.guild_id = str(guild_id or "").strip()
        self.submit_fn = submit_fn # Korrigiert
        self.log_fn = log_fn
        self.stop_event = threading.Event()
        self.loop: Optional[asyncio.AbstractEventLoop] = None
        self._client: Any = None
        self.workspace = _workspace_root(workspace)
        self.presence_cfg = dict(presence_cfg or {})
        self.audio_service = audio_service
        self.tts_cfg = tts_cfg or {}
        self.tts_enabled = bool(self.tts_cfg.get("enabled", False))
        self.tts_voice = str(self.tts_cfg.get("voice", "de-DE-ConradNeural"))
        self.reply_timeout_sec = max(10, min(3600, int(reply_timeout_sec or 600)))
        self._last_id_file = (self.workspace.parent / "sessions" / "discord_last_msg_id.txt")


    def stop(self) -> None:
        self.stop_event.set()
        if self.loop and self._client:
            with contextlib.suppress(Exception):
                asyncio.run_coroutine_threadsafe(self._client.close(), self.loop)

    def run(self) -> None:
        try:
            import discord  # type: ignore
            from discord import app_commands  # type: ignore
        except Exception as exc:
            self.log_fn(f"Discord Gateway nicht verfuegbar (discord.py fehlt): {exc}")
            return

        self.loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self.loop)
        self.log_fn("Discord Gateway Worker gestartet.")
        restart_delay = 3.0
        while not self.stop_event.is_set():
            intents = discord.Intents.default()
            intents.message_content = True
            intents.messages = True
            intents.guilds = True
            client = discord.Client(intents=intents)
            self._client = client
            tree = app_commands.CommandTree(client)

            channel_filter = str(self.channel_id or "").strip()
            parts = [p.strip() for p in channel_filter.split("/") if p.strip()]
            if len(parts) > 1:
                channel_filter = parts[-1]

            async def _reply_to_interaction(interaction: Any, text: str) -> None:
                msg = str(text or "").strip() or "Keine Antwort."
                chunks = [msg[i : i + 1900] for i in range(0, len(msg), 1900)] or ["Keine Antwort."]
                if not interaction.response.is_done():
                    await interaction.response.defer(ephemeral=True, thinking=False)
                for ch in chunks:
                    await interaction.followup.send(ch, ephemeral=True)

            async def _call_reply_async(text: str, meta: Dict[str, Any], platform: str) -> str:
                loop = asyncio.get_running_loop()
                meta_local = dict(meta or {})
                meta_local["trace_id"] = str(meta_local.get("trace_id") or uuid.uuid4().hex[:16])
                meta_local["reply_timeout_sec"] = self.reply_timeout_sec
                return await loop.run_in_executor(
                    None,
                    lambda: _run_reply_with_timeout(
                        self.submit_fn,
                        text,
                        meta_local,
                        timeout_sec=self.reply_timeout_sec,
                        log_fn=self.log_fn,
                        platform=platform,
                    ),
                )

            @tree.command(name="status", description="Zeigt K.AI Agent Status")
            async def slash_status(interaction: Any) -> None:  # type: ignore[misc]
                meta = {
                    "platform": "discord",
                    "channel_id": str(getattr(getattr(interaction, "channel", None), "id", "") or ""),
                    "user": str(getattr(getattr(interaction, "user", None), "name", "unknown") or "unknown"),
                    "is_dm": getattr(interaction, "guild", None) is None,
                    "attachments": [],
                }
                reply = await _call_reply_async("/status", meta, "discord-gateway")
                await _reply_to_interaction(interaction, str(reply or "Status nicht verfuegbar."))

            @tree.command(name="commands", description="Zeigt Befehlsuebersicht")
            async def slash_commands(interaction: Any) -> None:  # type: ignore[misc]
                meta = {
                    "platform": "discord",
                    "channel_id": str(getattr(getattr(interaction, "channel", None), "id", "") or ""),
                    "user": str(getattr(getattr(interaction, "user", None), "name", "unknown") or "unknown"),
                    "is_dm": getattr(interaction, "guild", None) is None,
                    "attachments": [],
                }
                reply = await _call_reply_async("/befehle", meta, "discord-gateway")
                await _reply_to_interaction(interaction, str(reply or "Keine Befehle verfuegbar."))

            gedaechtnis_group = app_commands.Group(name="gedaechtnis", description="Gedaechtnis Befehle")

            @gedaechtnis_group.command(name="anzeigen", description="Zeigt den Gedaechtnis-Status")
            async def slash_mem_show(interaction: Any) -> None:  # type: ignore[misc]
                meta = {
                    "platform": "discord",
                    "channel_id": str(getattr(getattr(interaction, "channel", None), "id", "") or ""),
                    "user": str(getattr(getattr(interaction, "user", None), "name", "unknown") or "unknown"),
                    "is_dm": getattr(interaction, "guild", None) is None,
                    "attachments": [],
                }
                reply = await _call_reply_async("/gedaechtnis anzeigen", meta, "discord-gateway")
                await _reply_to_interaction(interaction, str(reply or "Keine Antwort."))

            @gedaechtnis_group.command(name="letzte", description="Zeigt letzte Gedaechtnis-Eintraege")
            async def slash_mem_recent(interaction: Any, anzahl: int = 5) -> None:  # type: ignore[misc]
                meta = {
                    "platform": "discord",
                    "channel_id": str(getattr(getattr(interaction, "channel", None), "id", "") or ""),
                    "user": str(getattr(getattr(interaction, "user", None), "name", "unknown") or "unknown"),
                    "is_dm": getattr(interaction, "guild", None) is None,
                    "attachments": [],
                }
                n = max(1, min(20, int(anzahl or 5)))
                cmd = f"/gedaechtnis letzte eintraege {n}"
                reply = await _call_reply_async(cmd, meta, "discord-gateway")
                await _reply_to_interaction(interaction, str(reply or "Keine Antwort."))

            tree.add_command(gedaechtnis_group)

            @client.event
            async def on_ready() -> None:
                name = getattr(client.user, "name", "unknown")
                self.log_fn(f"Discord Gateway online als {name}.")
                try:
                    enabled = bool(self.presence_cfg.get("enabled", True))
                    if enabled:
                        status_raw = str(self.presence_cfg.get("status", "online") or "online").strip().lower()
                        status_map = {
                            "online": discord.Status.online,
                            "idle": discord.Status.idle,
                            "dnd": discord.Status.dnd,
                            "do_not_disturb": discord.Status.dnd,
                            "invisible": discord.Status.invisible,
                            "offline": discord.Status.invisible,
                        }
                        dc_status = status_map.get(status_raw, discord.Status.online)

                        activity_text = str(self.presence_cfg.get("text", "K.AI Agent aktiv") or "").strip()
                        activity_type_raw = str(self.presence_cfg.get("type", "watching") or "watching").strip().lower()
                        activity: Optional[Any] = None
                        if activity_text:
                            if activity_type_raw == "playing":
                                activity = discord.Game(name=activity_text)
                            elif activity_type_raw == "listening":
                                activity = discord.Activity(type=discord.ActivityType.listening, name=activity_text)
                            elif activity_type_raw == "competing":
                                activity = discord.Activity(type=discord.ActivityType.competing, name=activity_text)
                            else:
                                activity = discord.Activity(type=discord.ActivityType.watching, name=activity_text)
                        await client.change_presence(status=dc_status, activity=activity)
                except Exception as exc:
                    self.log_fn(f"Discord Presence setzen fehlgeschlagen: {exc}")
                try:
                    guild_obj = None
                    if self.guild_id.isdigit():
                        guild_obj = discord.Object(id=int(self.guild_id))
                    if guild_obj is not None:
                        tree.copy_global_to(guild=guild_obj)
                        synced = await tree.sync(guild=guild_obj)
                    else:
                        synced = await tree.sync()
                    self.log_fn(f"Discord Slash-Commands synchronisiert: {len(synced)}")
                except Exception as exc:
                    self.log_fn(f"Discord Slash-Command Sync fehlgeschlagen: {exc}")

            # Letzten bekannten Message-ID von Disk laden (verhindert Replay nach Neustart via RESUME)
            _seen_discord_msg_ids: set = set()
            _last_seen_snowflake: int = 0
            try:
                self._last_id_file.parent.mkdir(parents=True, exist_ok=True)
                if self._last_id_file.exists():
                    _saved = self._last_id_file.read_text(encoding="utf-8").strip()
                    if _saved.isdigit():
                        _last_seen_snowflake = int(_saved)
                        self.log_fn(f"Discord Gateway: Letzte bekannte Message-ID: {_saved}")
            except Exception:
                pass

            @client.event
            async def on_message(message: Any) -> None:
                nonlocal _last_seen_snowflake
                try:
                    if not message or not getattr(message, "author", None):
                        return
                    if bool(getattr(message.author, "bot", False)):
                        return
                    # Dedup: verhindert doppelte Events UND Replay nach Neustart (via RESUME)
                    _msg_id = str(getattr(message, "id", "") or "")
                    if _msg_id:
                        _mid_int = int(_msg_id) if _msg_id.isdigit() else 0
                        # Snowflake-Vergleich: Replay überspringen
                        if _mid_int and _mid_int <= _last_seen_snowflake:
                            return
                        if _msg_id in _seen_discord_msg_ids:
                            return
                        _seen_discord_msg_ids.add(_msg_id)
                        if len(_seen_discord_msg_ids) > 500:
                            _seen_discord_msg_ids.clear()
                        # Persistieren damit nach Neustart kein Replay passiert
                        if _mid_int > _last_seen_snowflake:
                            _last_seen_snowflake = _mid_int
                            try:
                                self._last_id_file.write_text(_msg_id, encoding="utf-8")
                            except Exception:
                                pass
                    msg_channel_id = str(getattr(message.channel, "id", ""))
                    msg_user = str(getattr(message.author, "name", "unknown") or "unknown")
                    is_dm = getattr(message, "guild", None) is None
                    if channel_filter and (not is_dm) and msg_channel_id != channel_filter:
                        return

                    text = str(getattr(message, "content", "") or "").strip()
                    attachments_paths: List[str] = []
                    atts = list(getattr(message, "attachments", []) or [])
                    if atts:
                        target_dir = _incoming_dir(self.workspace, "discord")
                        async with httpx.AsyncClient(timeout=60, trust_env=False) as ac:
                            for a in atts:
                                try:
                                    url = str(getattr(a, "url", "") or "")
                                    fname = _safe_name(str(getattr(a, "filename", "") or f"dc_{uuid.uuid4().hex[:8]}"))
                                    if not url:
                                        continue
                                    target = (target_dir / fname).resolve()
                                    rr = await ac.get(url)
                                    rr.raise_for_status()
                                    target.write_bytes(rr.content)
                                    attachments_paths.append(str(target))
                                except Exception as exc:
                                    self.log_fn(f"Discord Gateway attachment download Fehler: {exc}")

                    # SST Logic
                    transcriptions = []
                    user_sent_voice = False
                    if self.audio_service:
                        for ap in attachments_paths:
                             pp = Path(ap)
                             if pp.suffix.lower() in [".ogg", ".mp3", ".wav", ".m4a"]:
                                 user_sent_voice = True
                                 with contextlib.suppress(Exception):
                                     await message.channel.typing()
                                 tr = await self.loop.run_in_executor(None, self.audio_service.transcribe, pp)
                                 if tr and not tr.startswith("[SST Error"):
                                     transcriptions.append(tr)
                   
                    if transcriptions:
                        if text.strip():
                            text = (text + "\n\n" + "\n".join(transcriptions)).strip()
                        else:
                            text = "\n".join(transcriptions).strip()

                    # Bilder für Vision-Analyse enkodieren, Dokumente per Text-Extraktion
                    _image_blobs: List[Dict[str, Any]] = []
                    _non_image_paths: List[str] = []
                    _too_large_hints: List[str] = []
                    _doc_contents: List[str] = []
                    for _ap in attachments_paths:
                        _blob = _encode_image(_ap)
                        if _blob is None:
                            _doc_text = _extract_document_content(_ap)
                            if _doc_text:
                                _fname = Path(_ap).name
                                _doc_contents.append(f"📄 Inhalt von '{_fname}':\n{_doc_text}")
                            else:
                                _non_image_paths.append(_ap)
                        elif _blob.get("too_large"):
                            _too_large_hints.append(f"⚠️ Bild '{_blob['filename']}' ({_blob['size_mb']} MB) ist zu groß für Bildanalyse (max 3 MB) und wurde als Datei gespeichert.")
                            _non_image_paths.append(_ap)
                        else:
                            _image_blobs.append(_blob)

                    effective_text = _append_attachment_context(text, _non_image_paths)
                    if _doc_contents:
                        _doc_block = "\n\n".join(_doc_contents)
                        effective_text = (effective_text + "\n\n" + _doc_block).strip() if effective_text else _doc_block
                    if _too_large_hints:
                        effective_text = (effective_text + "\n\n" + "\n".join(_too_large_hints)).strip()
                    if not effective_text and not _image_blobs:
                        return
                    if not effective_text:
                        effective_text = "Bild erhalten."
                    with contextlib.suppress(Exception):
                        await message.channel.typing()
                    # --- Asynchrone Queue Einreichung (Discord Gateway) ---
                    async def _on_finish_dc_gw(res: Dict[str, Any]):
                        reply_text = _ensure_reply_text(res.get("reply", ""))
                        steps_gw = res.get("steps", [])
                        if reply_text:
                            try:
                                # Direkt awaiten — wir sind bereits im Event-Loop
                                for i in range(0, len(reply_text), 1900):
                                    await message.channel.send(reply_text[i:i+1900])
                            except Exception as _dc_send_err:
                                self.log_fn(f"Discord Gateway send Fehler: {_dc_send_err}")
                        agent_sent_file_gw = any(
                            isinstance(s, dict) and s.get("kind") == "send_messenger_file" and s.get("ok")
                            for s in steps_gw
                        )
                        if not agent_sent_file_gw:
                            for p in _extract_paths_for_send(reply_text, workspace=self.workspace):
                                try:
                                    import discord as _dc_mod
                                    await message.channel.send(file=_dc_mod.File(str(p)))
                                except Exception:
                                    pass

                    # In Queue schieben
                    _dc_submit_kwargs: Dict[str, Any] = dict(
                        result_callback=lambda r: asyncio.run_coroutine_threadsafe(_on_finish_dc_gw(r), self.loop) if self.loop else None,
                        source="messenger:discord-gateway",
                        dialog_key=f"dc:{msg_channel_id}",
                        platform="discord",
                        channel_id=msg_channel_id,
                        user=msg_user,
                    )
                    if _image_blobs:
                        _dc_submit_kwargs["images"] = _image_blobs
                    self.submit_fn(effective_text, **_dc_submit_kwargs)
                    # Optisches Feedback
                    with contextlib.suppress(Exception):
                        await message.add_reaction("⏳")
                except Exception as exc:
                    self.log_fn(f"Discord Gateway on_message Fehler: {exc}")

            async def runner() -> None:
                start_task = asyncio.create_task(client.start(self.token))
                try:
                    while not self.stop_event.is_set():
                        if start_task.done():
                            with contextlib.suppress(Exception):
                                exc = start_task.exception()
                                if exc:
                                    self.log_fn(f"Discord Gateway client task beendet: {exc}")
                            break
                        await asyncio.sleep(0.5)
                finally:
                    with contextlib.suppress(Exception):
                        await client.close()
                    with contextlib.suppress(Exception):
                        await start_task

            try:
                self.loop.run_until_complete(runner())
            except Exception as exc:
                self.log_fn(f"Discord Gateway Fehler: {exc}")
            if not self.stop_event.is_set():
                self.log_fn(f"Discord Gateway Restart in {int(restart_delay)}s...")
                time.sleep(restart_delay)

        with contextlib.suppress(Exception):
            if self.loop and not self.loop.is_closed():
                self.loop.stop()
                self.loop.close()
        self.log_fn("Discord Gateway Worker gestoppt.")


class MessengerRuntime:
    def __init__(self):
        self.telegram: Optional[TelegramWorker] = None
        self.discord: Optional[threading.Thread] = None

    def start(self, cfg: Dict[str, Any], reply_fn: ReplyFn, log_fn: LogFn) -> None:
        msg_cfg = cfg.get("messenger", {})
        tts_cfg = cfg.get("tts", {})
        tg = msg_cfg.get("telegram", {})
        dc = msg_cfg.get("discord", {})
        reply_timeout_sec = max(10, min(3600, int(msg_cfg.get("reply_timeout_sec", 600) or 600)))
        ws_raw = str(cfg.get("workspace", "data/workspace"))
        workspace = Path(ws_raw)
        if not workspace.is_absolute():
            from pathlib import Path as P
            workspace = (P(__file__).parents[2] / ws_raw).resolve()

        audio_service = None
        try:
            # Only init if requested/needed to save resources on startup
            # We check if TTS is enabled or if we want SST (implicit for now)
            # Actually, let's always init it, it lazily loads models anyway.
            audio_service = AudioService(workspace)
        except Exception as exc:
            log_fn(f"Audio Service konnte nicht initialisiert werden: {exc}")

        tg_enabled = bool(tg.get("enabled"))
        tg_token = str(tg.get("token", "") or "")
        if tg_enabled and tg_token:
            self.telegram = TelegramWorker(
                token=tg_token, 
                submit_fn=reply_fn, # In main.py ist reply_fn bereits die Queue-Submit-kompatible Funktion
                log_fn=log_fn, 
                workspace=workspace,
                audio_service=audio_service,
                tts_cfg=tts_cfg,
                reply_timeout_sec=reply_timeout_sec,
            )
            self.telegram.start()
        else:
            log_fn("Telegram Worker nicht gestartet (disabled oder token fehlt).")

        dc_enabled = bool(dc.get("enabled"))
        dc_token = str(dc.get("token", "") or "")
        dc_channel_id = str(dc.get("channel_id", "") or "")
        dc_guild_id = ""
        dc_gateway_enabled = bool(dc.get("gateway_enabled", True))
        dc_presence_cfg = dc.get("presence", {}) if isinstance(dc.get("presence", {}), dict) else {}
        if dc_channel_id:
            parts = [p.strip() for p in dc_channel_id.split("/") if p.strip()]
            if len(parts) > 1:
                dc_guild_id = parts[0]
                dc_channel_id = parts[-1]
        if dc_enabled and dc_token and dc_gateway_enabled:
            self.discord = DiscordGatewayWorker(
                token=dc_token,
                channel_id=dc_channel_id,
                guild_id=dc_guild_id,
                submit_fn=reply_fn, # Geändert
                log_fn=log_fn,
                workspace=workspace,
                presence_cfg=dc_presence_cfg,
                audio_service=audio_service,
                tts_cfg=tts_cfg,
                reply_timeout_sec=reply_timeout_sec,
            )
            self.discord.start()
        elif dc_enabled and dc_token and dc_channel_id:
            self.discord = DiscordRestWorker(
                token=dc_token,
                channel_id=dc_channel_id,
                submit_fn=reply_fn, # Geändert
                log_fn=log_fn,
                workspace=workspace,
                audio_service=audio_service,
                tts_cfg=tts_cfg,
                reply_timeout_sec=reply_timeout_sec,
            )
            self.discord.start()
        elif dc_enabled and dc_token and not dc_channel_id:
            log_fn("Discord REST Worker nicht gestartet (channel_id fehlt). Fuer Online-Status Gateway aktivieren.")
        else:
            log_fn("Discord Worker nicht gestartet (disabled oder token fehlt).")

    def stop(self) -> None:
        if self.telegram:
            self.telegram.stop()
            self.telegram = None
        if self.discord:
            stop_fn = getattr(self.discord, "stop", None)
            if callable(stop_fn):
                stop_fn()
            self.discord = None
